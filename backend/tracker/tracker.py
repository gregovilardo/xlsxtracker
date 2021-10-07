from openpyxl import load_workbook
from openpyxl.workbook import workbook

def track_coins(filename):

    wb = load_workbook(filename=filename, data_only=True)
    sheet = wb.active

    trades = {}


    excel = list(sheet.iter_rows(min_row=2,
                                    # max_row=33,
                                    max_col=7,
                                    values_only=True))
    counter = len(excel)-1

    """here all trades are saved in a variable with fees already descount"""

    for value in excel:
        if value[2] == "SELL":
            trade = {
                "date": value[0],
                "pair": value[1],
                "type": value[2],
                "price": float(value[3]),
                "ammount": float(value[4]),
                "total": float(value[5]) - float(value[6]),
                "trade_fullfill": False,
            }
        else:
            trade = {
                "date": value[0],
                "pair": value[1],
                "type": value[2],
                "price": float(value[3]),
                "ammount": float(value[4]) - float(value[6]),
                "total": float(value[5]),
                "trade_fullfill": False,
            }
            
        trades[counter] = trade
        counter -= 1


    profit_loss = []
    list_to_fullfill = []

    for i in range(len(trades)):
        if trades[i]["type"] == "SELL" or trades[i]["trade_fullfill"] == True:
            continue
        total_ammount = trades[i]["ammount"]
        total_buy = trades[i]["total"]
        total_sell = 0
        for j in range(i + 1, len(trades)):
            if trades[i]["pair"] == trades[j]["pair"]:
                if trades[j]["type"] == "SELL":
                    total_ammount -= trades[j]["ammount"]
                    total_sell += trades[j]["total"]
                else:
                    total_ammount += + trades[j]["ammount"]
                    total_buy += + trades[j]["total"]
                avg_resto = 1 / trades[j]["price"]
                list_to_fullfill.append(j)
                if total_ammount > -avg_resto and total_ammount < avg_resto:
                    trades[i]["trade_fullfill"] = True
                    for n in range(len(list_to_fullfill)):
                        trades[list_to_fullfill[n]]["trade_fullfill"] = True
                        # print(json.dumps(trades[n], indent=4))
                        # print(n)
                    list_to_fullfill = []
                    total_total = total_sell - total_buy
                    finish_trade = {
                        "pair": trades[i]["pair"],
                        "start": trades[j]["date"],
                        "end": trades[j]["date"],
                        "profit_loss": total_total,
                        "percentage": (trades[j]["price"] * 100 / trades[i]["price"]) - 100,
                    }
                    profit_loss.append(finish_trade)
                    break

    holds = {}
    ctr = -1
    trades_on_hold = []

    for i in range(len(trades)):
        if trades[i]["trade_fullfill"] == False:
            if trades[i]["type"] == "BUY":
                ctr += 1
                hold = {
                    "pair": trades[i]["pair"],
                    "start_date": trades[i]["date"],
                    "holding": trades[i]["ammount"],
                    "start_price": trades[i]["price"],
                    "start_total": trades[i]["total"],
                }
                holds[ctr] = hold
                for j in range(i + 1, len(trades)):
                    if trades[i]["pair"] == trades[j]["pair"]:
                        if trades[j]["type"] == "BUY":
                            holds[ctr]["holding"] += trades[j]["ammount"]
                        else:
                            holds[ctr]["holding"] -= trades[j]["ammount"]

            else:
                trades_on_hold.append(trades[i])


    for trade in trades_on_hold:
        for i in range(len(holds)):
            if trade["pair"] == holds[i]["pair"]:
                trade["hold"] = holds[i]
                trade["profit_loss"] = -(holds[i]["start_price"] * trade["ammount"]) + trade["total"]
                trade["percentage"] = (trade["price"] * 100 / holds[i]["start_price"]) -100

    holds_list = []
    for i in range(len(holds)):
        holds_list.append(holds[i])
            
    print(holds)
    return {"trades":profit_loss, "holds":holds_list, "trades_on_hold":trades_on_hold}