from openpyxl import load_workbook
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(filename="sheet.xlsx", data_only=True)
sheet = wb.active
wb2 = load_workbook(filename="sheet2.xlsx", data_only=True)
sheet2 = wb2.active

wb = Workbook()

dest_filename = 'TradingHistory.xlsx'

ws = wb.active
ws.title = "Trades"


for row in range(len(list(sheet))):
    for col in range(7):
        print(sheet(col, row))


# excel = list(sheet.iter_rows(min_row=2,
#                                 # max_row=33,
#                                 max_col=7,
#                                 values_only=True))

# for value in excel:
#     wb[counter]

# ws['A1'] = "HOLA GREGO"

# wb.save(filename = dest_filename)





